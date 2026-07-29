"""把统一实验结果导出为 Excel。"""

from pathlib import Path
from openpyxl import Workbook

from core.schemas import ExperimentResult


def export_experiment_result_to_excel(
    experiment_result: ExperimentResult,
    output_path: str,
) -> str:
    """生成三个工作表的 xlsx 文件，并返回文件路径。"""
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "整理后数据"

    headers = [column.internal_name for column in experiment_result.columns]
    data_sheet.append(["序号", *headers])
    included_rows = [row for row in experiment_result.rows if row.include_in_final]
    for serial_number, row in enumerate(included_rows, start=1):
        data_sheet.append([serial_number, *[row.values.get(header, "") for header in headers]])

    field_sheet = workbook.create_sheet("字段说明")
    field_sheet.append([
        "字段内部名称", "最终中文名称", "原始单位", "AI建议单位", "最终确认单位",
        "字段来源", "是否经过用户确认",
    ])
    for column in experiment_result.columns:
        field_sheet.append([
            column.internal_name,
            column.display_name,
            column.original_unit or column.unit or "",
            column.suggested_unit or "",
            column.final_unit or "",
            column.source,
            "是" if experiment_result.user_confirmed else "否",
        ])

    record_sheet = workbook.create_sheet("识别记录")
    record_sheet.append([
        "原始文件名", "原始识别文本", "AI 建议结果", "AI 提供商", "不确定内容", "警告",
        "修改记录", "最终值来源",
    ])
    record_sheet.append([
        "、".join(file.filename for file in experiment_result.source_files),
        experiment_result.raw_text,
        str([row.values for row in experiment_result.ai_suggested_rows]),
        experiment_result.provider,
        "；".join(item.content for item in experiment_result.uncertain_items),
        "；".join(experiment_result.warnings),
        "；".join(experiment_result.modification_records),
        str([
            {
                "row": index + 1,
                "include_in_final": row.include_in_final,
                "observed_values": row.observed_values,
                "sources": row.field_sources,
            }
            for index, row in enumerate(experiment_result.rows)
        ]),
    ])

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return str(path)
