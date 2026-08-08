"""AI 科研数据助手的可重复图片基准测试入口。

默认使用 Mock Provider，避免无意产生 API 费用。传入
``--provider doubao --validation-provider doubao`` 才会调用真实模型。
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import tempfile
import time
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any


PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT))

from core.excel_exporter import export_experiment_result_to_excel
from core.experiment_parser import process_experiment_images
from core.schemas import ExperimentResult, SourceFile
from core.validation_service import validate_experiment_result


IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg"}
VISUAL_FIELD_PATTERN = re.compile(
    r"^(?:col\d*|column\d*|column_group|column_label|layout_column|"
    r"left(?:_column)?|middle(?:_column)?|right(?:_column)?|desc|val|field\d*)$",
    re.IGNORECASE,
)
NUMBER_WITH_LETTER_PATTERN = re.compile(r"[UVEO]", re.IGNORECASE)


def source_file_from_path(image_path: Path) -> SourceFile:
    """读取一张本地图片，转换为项目的统一输入对象。"""
    suffix = image_path.suffix.lower()
    if suffix not in IMAGE_SUFFIXES:
        raise ValueError(f"不支持的图片格式：{image_path.name}")
    file_type = "image/png" if suffix == ".png" else "image/jpeg"
    return SourceFile(image_path.name, file_type, image_path.read_bytes())


def identifier_symbol_anomalies(result: ExperimentResult) -> list[dict[str, Any]]:
    """仅标记可明确识别的可疑编号模式，不自动修复任何字符。"""
    patterns = {
        "vertical_bar": re.compile(r"\|"),
        "slash_segment_as_1_i_l": re.compile(r"/[1Il](?=/)"),
        "missing_slash_before_e": re.compile(r"(?:L|H)[1Il]E[+-]"),
        "missing_slash_before_n": re.compile(r"E[+-][1Il]N"),
        "merged_identifier_fragments": re.compile(r"LIE|HIE|E[+-]I?N\d"),
    }
    anomalies: list[dict[str, Any]] = []
    for row_number, row in enumerate(result.rows, start=1):
        for field_name in ("treatment_id", "sample_id"):
            value = row.values.get(field_name)
            if not isinstance(value, str):
                continue
            matched = [name for name, pattern in patterns.items() if pattern.search(value)]
            if matched:
                anomalies.append({
                    "row": row_number,
                    "field": field_name,
                    "value": value,
                    "patterns": matched,
                })
    return anomalies


def numeric_letter_errors(result: ExperimentResult) -> list[dict[str, Any]]:
    """记录数值语义字段中可能混入的字母；保持模型原始值不变。"""
    errors: list[dict[str, Any]] = []
    for row_number, row in enumerate(result.rows, start=1):
        for field_name, value in row.values.items():
            is_numeric_field = any(token in field_name.lower() for token in (
                "measurement", "value", "mass", "reading", "result",
            ))
            if is_numeric_field and isinstance(value, str) and NUMBER_WITH_LETTER_PATTERN.search(value):
                errors.append({"row": row_number, "field": field_name, "value": value})
    return errors


def identifier_field_summary(result: ExperimentResult) -> dict[str, int]:
    """统计 treatment_id/sample_id 的使用情况，不判断其真实内容是否正确。"""
    treatment_rows = 0
    sample_rows = 0
    missing_identifier_rows = 0
    for row in result.rows:
        has_treatment = row.values.get("treatment_id") is not None
        has_sample = row.values.get("sample_id") is not None
        treatment_rows += int(has_treatment)
        sample_rows += int(has_sample)
        missing_identifier_rows += int(not has_treatment and not has_sample)
    return {
        "treatment_id_rows": treatment_rows,
        "sample_id_rows": sample_rows,
        "missing_identifier_rows": missing_identifier_rows,
    }


def replicate_summary(result: ExperimentResult) -> dict[str, Any]:
    """检查重复测量是否被压缩；不判断图片真实答案。"""
    compressed_values: list[dict[str, Any]] = []
    plural_fields: list[str] = []
    indexed_rows: list[dict[str, Any]] = []
    groups: dict[str, list[int]] = {}

    for row_number, row in enumerate(result.rows, start=1):
        if row.replicate_index is not None:
            indexed_rows.append({
                "row": row_number,
                "replicate_group": row.replicate_group,
                "replicate_index": row.replicate_index,
            })
            if row.replicate_group:
                groups.setdefault(row.replicate_group, []).append(row.replicate_index)
        for field_name, value in row.values.items():
            if field_name.lower().endswith("measurements"):
                plural_fields.append(field_name)
            if isinstance(value, (list, tuple, dict)):
                compressed_values.append({"row": row_number, "field": field_name, "kind": type(value).__name__})
            elif "measurement" in field_name.lower() and isinstance(value, str):
                if re.search(r"\d\s*,\s*\d", value):
                    compressed_values.append({"row": row_number, "field": field_name, "kind": "comma_text"})

    nonsequential_groups = {
        group: indexes
        for group, indexes in groups.items()
        if sorted(indexes) != list(range(1, len(indexes) + 1))
    }
    return {
        "compressed_values": compressed_values,
        "plural_measurement_fields": sorted(set(plural_fields)),
        "replicate_indexed_rows": indexed_rows,
        "nonsequential_groups": nonsequential_groups,
    }


def review_ready_count(result: ExperimentResult, validation_result: Any) -> int:
    """统计可定位、可进入人工确认区的单元格数量。"""
    keys: set[tuple[int, str]] = set()
    columns = {column.internal_name for column in result.columns}
    for finding in [*validation_result.suggestions, *validation_result.uncertain_items]:
        if (
            finding.scope == "cell"
            and isinstance(finding.row_index, int)
            and 1 <= finding.row_index <= len(result.rows)
            and finding.column_name in columns
        ):
            keys.add((finding.row_index, finding.column_name))
    return len(keys) + len(result.uncertain_items)


def export_check(result: ExperimentResult) -> dict[str, Any]:
    """在系统临时目录生成 Excel，确认导出函数可运行且包含三张工作表。"""
    try:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory(prefix="scientific-benchmark-") as directory:
            output_path = Path(directory) / "result.xlsx"
            export_experiment_result_to_excel(result, str(output_path))
            workbook = load_workbook(output_path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
        return {
            "success": True,
            "sheet_names": sheet_names,
            "has_three_sheets": len(sheet_names) == 3,
        }
    except Exception as error:  # 基准测试需要记录失败，而不是中止整个批次。
        return {"success": False, "error": f"{type(error).__name__}: {error}"}


def benchmark_one_image(
    image_path: Path,
    provider_name: str,
    validation_provider_name: str,
    enable_preprocessing: bool = False,
) -> dict[str, Any]:
    """运行一张图片的 Extraction、Validation 与 Excel 检查。"""
    record: dict[str, Any] = {
        "filename": image_path.name,
        "file_size_bytes": image_path.stat().st_size,
        "extraction": {"success": False},
        "validation": {"success": False},
        "review": {},
        "excel": {},
    }
    try:
        source = source_file_from_path(image_path)
        extraction_started = time.perf_counter()
        result = process_experiment_images(
            [source],
            provider_name=provider_name,
            enable_preprocessing=enable_preprocessing,
        )
        extraction_seconds = round(time.perf_counter() - extraction_started, 3)
        visual_fields = [
            column.internal_name
            for column in result.columns
            if VISUAL_FIELD_PATTERN.match(column.internal_name)
        ]
        split_rows = [
            row_number
            for row_number, row in enumerate(result.rows, start=1)
            if "sample_id" in row.values and "experiment_description" in row.values
        ]
        record["extraction"] = {
            "success": True,
            "api_seconds": extraction_seconds,
            "row_count": len(result.rows),
            "columns": [column.internal_name for column in result.columns],
            "identifier_fields": identifier_field_summary(result),
            "visual_fields": visual_fields,
            "treatment_id_split_rows": split_rows,
            "identifier_symbol_anomalies": identifier_symbol_anomalies(result),
            "numeric_letter_errors": numeric_letter_errors(result),
            "replicate": replicate_summary(result),
        }
    except Exception as error:
        record["extraction"] = {
            "success": False,
            "error": f"{type(error).__name__}: {error}",
        }
        return record

    try:
        validation_started = time.perf_counter()
        validation_result = validate_experiment_result(result, validation_provider_name)
        record["validation"] = {
            "success": True,
            "api_seconds": round(time.perf_counter() - validation_started, 3),
            "warnings_count": len(validation_result.warnings),
            "suggestions_count": len(validation_result.suggestions),
            "uncertain_items_count": len(validation_result.uncertain_items),
        }
        record["review"] = {
            "extraction_uncertain_items_count": len(result.uncertain_items),
            "validation_review_ready_count": review_ready_count(result, validation_result),
        }
    except Exception as error:
        record["validation"] = {
            "success": False,
            "error": f"{type(error).__name__}: {error}",
        }
    record["excel"] = export_check(result)
    return record


def build_summary(report: dict[str, Any]) -> dict[str, Any]:
    """从逐图记录计算长期回归测试使用的稳定汇总指标。"""
    images = report["images"]
    total = len(images)
    extraction_successes = [item for item in images if item["extraction"].get("success")]
    validation_successes = [item for item in images if item["validation"].get("success")]
    excel_successes = [item for item in images if item["excel"].get("success")]

    def average(values: list[float]) -> float | None:
        return round(sum(values) / len(values), 3) if values else None

    return {
        "total_images": total,
        "extraction_success_count": len(extraction_successes),
        "extraction_success_rate": round(len(extraction_successes) / total, 3) if total else 0.0,
        "validation_success_count": len(validation_successes),
        "validation_success_rate": round(len(validation_successes) / total, 3) if total else 0.0,
        "excel_success_count": len(excel_successes),
        "excel_success_rate": round(len(excel_successes) / total, 3) if total else 0.0,
        "average_extraction_seconds": average([
            item["extraction"].get("api_seconds", 0.0) for item in extraction_successes
        ]),
        "average_row_count": average([
            float(item["extraction"].get("row_count", 0)) for item in extraction_successes
        ]),
        "field_anomaly_count": sum(
            len(item["extraction"].get("visual_fields", []))
            + len(item["extraction"].get("treatment_id_split_rows", []))
            for item in extraction_successes
        ),
        "identifier_symbol_anomaly_count": sum(
            len(item["extraction"].get("identifier_symbol_anomalies", []))
            for item in extraction_successes
        ),
        "numeric_letter_error_count": sum(
            len(item["extraction"].get("numeric_letter_errors", []))
            for item in extraction_successes
        ),
        "compressed_replicate_value_count": sum(
            len(item["extraction"].get("replicate", {}).get("compressed_values", []))
            for item in extraction_successes
        ),
        "review_ready_total": sum(
            item["review"].get("validation_review_ready_count", 0)
            for item in images
        ),
    }


def read_image_paths(arguments: argparse.Namespace) -> list[Path]:
    """读取命令行图片与可选清单文件，去重后保持输入顺序。"""
    raw_paths = list(arguments.images or [])
    if arguments.images_file:
        for line in Path(arguments.images_file).read_text(encoding="utf-8").splitlines():
            cleaned = line.strip()
            if cleaned and not cleaned.startswith("#"):
                raw_paths.append(cleaned)
    unique: list[Path] = []
    seen: set[Path] = set()
    for raw_path in raw_paths:
        path = Path(raw_path).expanduser().resolve()
        if path not in seen:
            unique.append(path)
            seen.add(path)
    if not unique:
        raise ValueError("请至少提供一张图片路径或 --images-file。")
    return unique


def report_markdown(report: dict[str, Any]) -> str:
    """生成适合人工查看的简短 Markdown 摘要。"""
    lines = [
        "# 图片基准测试报告",
        "",
        f"- 生成时间：{report['generated_at']}",
        f"- Extraction Provider：{report['provider']}",
        f"- Validation Provider：{report['validation_provider']}",
        f"- 总图片数：{report['summary']['total_images']}",
        f"- Extraction 成功率：{report['summary']['extraction_success_rate']:.1%}",
        f"- Validation 成功率：{report['summary']['validation_success_rate']:.1%}",
        f"- Excel 成功率：{report['summary']['excel_success_rate']:.1%}",
        f"- 平均 Extraction 耗时：{report['summary']['average_extraction_seconds']} 秒",
        f"- 平均识别行数：{report['summary']['average_row_count']}",
        f"- 字段异常次数：{report['summary']['field_anomaly_count']}",
        f"- 编号字符异常次数：{report['summary']['identifier_symbol_anomaly_count']}",
        "",
        "| 图片 | Extraction | 行数 | Validation findings | Excel |",
        "|---|---:|---:|---:|---:|",
    ]
    for item in report["images"]:
        extraction = item["extraction"]
        validation = item["validation"]
        findings = (
            validation.get("suggestions_count", 0)
            + validation.get("uncertain_items_count", 0)
        )
        lines.append(
            f"| {item['filename']} | {'成功' if extraction['success'] else '失败'} "
            f"| {extraction.get('row_count', '-')} | {findings if validation['success'] else '-'} "
            f"| {'成功' if item['excel'].get('success') else '失败'} |"
        )
    lines.extend(["", "详细指标请查看同名 JSON 报告。"])
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description="运行 AI 科研数据助手图片基准测试")
    parser.add_argument("images", nargs="*", help="一张或多张本地图片路径")
    parser.add_argument("--images-file", help="UTF-8 文本清单：每行一个图片路径")
    parser.add_argument("--provider", choices=["mock", "doubao"], default="mock")
    parser.add_argument("--validation-provider", choices=["mock", "doubao"], default="mock")
    parser.add_argument("--enable-preprocessing", action="store_true")
    parser.add_argument(
        "--output-dir",
        default=str(Path(__file__).parent / "results"),
        help="JSON 与 Markdown 报告目录",
    )
    arguments = parser.parse_args()
    paths = read_image_paths(arguments)

    results = []
    for path in paths:
        if not path.is_file():
            results.append({
                "filename": path.name,
                "extraction": {"success": False, "error": "图片文件不存在"},
                "validation": {"success": False},
                "review": {},
                "excel": {},
            })
            continue
        results.append(benchmark_one_image(
            path,
            arguments.provider,
            arguments.validation_provider,
            arguments.enable_preprocessing,
        ))

    report = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "provider": arguments.provider,
        "validation_provider": arguments.validation_provider,
        "images": results,
    }
    report["summary"] = build_summary(report)
    output_dir = Path(arguments.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "benchmark_report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8",
    )
    (output_dir / "benchmark_report.md").write_text(
        report_markdown(report), encoding="utf-8",
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))
    print(f"报告目录：{output_dir.resolve()}")


if __name__ == "__main__":
    main()
